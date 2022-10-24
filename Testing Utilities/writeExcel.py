import openpyxl

workbook = openpyxl.load_workbook(".//excel//testdata.xlsx")
sheet = workbook["LoginTest"]
totalrows = sheet.max_row
totalcols = sheet.max_column


# sheet.cell(row=1,column=3).value="Age"
# workbook.save(".//excel//testdata.xlsx")

for rows in range(4,8):
    for columns in range(1,4):
        sheet.cell(row=rows,column=columns).value = "####"

workbook.save(".//excel//testdata.xlsx")